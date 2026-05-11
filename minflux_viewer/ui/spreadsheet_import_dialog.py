from __future__ import annotations

import csv
from pathlib import Path

import numpy as np
from PyQt6.QtWidgets import (
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from ..core.dataset import build_localization_dataset


def _read_spreadsheet(path: str) -> tuple[list[str], list[dict[str, object]]]:
    file_path = Path(path)
    ext = file_path.suffix.lower()
    if ext in {".csv", ".tsv"}:
        delimiter = "\t" if ext == ".tsv" else ","
        with open(file_path, newline="", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh, delimiter=delimiter)
            if reader.fieldnames is None:
                raise ValueError("Spreadsheet has no header row.")
            rows = list(reader)
            return list(reader.fieldnames), rows
    if ext in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ImportError("openpyxl is required to open Excel spreadsheets.") from exc
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        values = list(ws.iter_rows(values_only=True))
        if not values:
            raise ValueError("Spreadsheet is empty.")
        headers = [str(v).strip() if v is not None else f"Column{i+1}" for i, v in enumerate(values[0])]
        rows = []
        for row in values[1:]:
            rows.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
        return headers, rows
    raise ValueError(f"Unsupported spreadsheet format: {ext}")


def _numeric_columns(headers: list[str], rows: list[dict[str, object]]) -> dict[str, np.ndarray]:
    out: dict[str, np.ndarray] = {}
    for header in headers:
        vals = []
        ok = True
        for row in rows:
            value = row.get(header, None)
            if value in ("", None):
                vals.append(np.nan)
                continue
            try:
                vals.append(float(value))
            except Exception:
                ok = False
                break
        if ok:
            out[header] = np.asarray(vals, dtype=float)
    return out


class SpreadsheetImportDialog(QDialog):
    ATTRIBUTE_FIELDS = ["xnm", "ynm", "znm", "tid", "tim"]

    def __init__(self, path: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._path = path
        self._headers, self._rows = _read_spreadsheet(path)
        self._numeric = _numeric_columns(self._headers, self._rows)
        if not self._numeric:
            raise ValueError("No numeric columns were found in the spreadsheet.")
        self._combos: dict[str, QComboBox] = {}

        self.setWindowTitle(f"Open spreadsheet data — {Path(path).name}")
        self.resize(900, 560)
        self._build_ui()

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.addWidget(QLabel("Assign spreadsheet columns to viewer attributes. `xnm` and `ynm` are required."))

        form = QFormLayout()
        options = ["<None>"] + list(self._numeric.keys())
        defaults = {
            "xnm": ["xnm", "x_nm", "x"],
            "ynm": ["ynm", "y_nm", "y"],
            "znm": ["znm", "z_nm", "z"],
            "tid": ["tid", "trace_id"],
            "tim": ["tim", "time"],
        }
        for field in self.ATTRIBUTE_FIELDS:
            combo = QComboBox()
            combo.addItems(options)
            for candidate in defaults.get(field, []):
                idx = combo.findText(candidate)
                if idx >= 0:
                    combo.setCurrentIndex(idx)
                    break
            form.addRow(field, combo)
            self._combos[field] = combo
        root.addLayout(form)

        root.addWidget(QLabel("Preview"))
        table = QTableWidget()
        preview_rows = min(8, len(self._rows))
        table.setRowCount(preview_rows)
        table.setColumnCount(len(self._headers))
        table.setHorizontalHeaderLabels(self._headers)
        for r in range(preview_rows):
            for c, header in enumerate(self._headers):
                table.setItem(r, c, QTableWidgetItem(str(self._rows[r].get(header, ""))))
        root.addWidget(table, stretch=1)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        root.addWidget(buttons)

    def build_dataset(self):
        selected = {field: combo.currentText() for field, combo in self._combos.items()}
        if selected["xnm"] == "<None>" or selected["ynm"] == "<None>":
            raise ValueError("Please assign both xnm and ynm columns.")

        x_nm = self._numeric[selected["xnm"]]
        y_nm = self._numeric[selected["ynm"]]
        z_nm = self._numeric[selected["znm"]] if selected["znm"] != "<None>" else None

        attrs = {}
        for field in ("tid", "tim"):
            if selected[field] != "<None>":
                attrs[field] = self._numeric[selected[field]]
        for header, arr in self._numeric.items():
            if header not in selected.values():
                attrs[header] = arr

        return build_localization_dataset(
            name=Path(self._path).name,
            folder=str(Path(self._path).parent),
            x_nm=x_nm,
            y_nm=y_nm,
            z_nm=z_nm,
            attrs=attrs,
        )


def choose_spreadsheet_and_import(parent=None):
    path, _ = QFileDialog.getOpenFileName(
        parent,
        "Open spreadsheet data",
        "",
        "Spreadsheet (*.csv *.tsv *.xlsx *.xlsm);;All files (*)",
    )
    if not path:
        return None
    dlg = SpreadsheetImportDialog(path, parent=parent)
    if dlg.exec() != QDialog.DialogCode.Accepted:
        return None
    return dlg.build_dataset()
