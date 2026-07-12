"""
minflux_viewer.core.spreadsheet_loader
=======================================
Generic spreadsheet localization importer (Phase 1 — headless core).

Reads delimited text (``.csv`` / ``.tsv`` / ``.txt``) and Excel (``.xlsx`` /
``.xlsm``) tables and maps their columns to the canonical
``loc_x`` / ``loc_y`` / ``loc_z`` (**metres**) model so imported data renders
and analyses like a native dataset.

It recognises the column conventions of common single-molecule-localization
tools so well-known exports map automatically:

* **ThunderSTORM** (Fiji) — ``x [nm]``, ``y [nm]``, ``z [nm]``,
  ``uncertainty[_xy/_z] [nm]``, ``id``, ``frame``, ``intensity [photon]`` (nm).
* **SMAP** (Ries lab) — ``xnm``, ``ynm``, ``znm``, ``locprecnm``,
  ``locprecznm``, ``groupindex``, ``frame``, ``phot`` (nm).
* **Picasso** (Jungmann lab) — ``x``, ``y``, ``z``, ``lpx``, ``lpy``,
  ``group``, ``frame``, ``photons`` (camera **pixels** → needs a pixel size).

The UI mapping dialog (Phase 2) drives this module; everything here is headless
and unit-tested.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path

import numpy as np

# ---------------------------------------------------------------------------
# Roles & units
# ---------------------------------------------------------------------------

#: The semantic roles a column can be mapped to. ``x`` and ``y`` are required.
ROLES: tuple[str, ...] = ("x", "y", "z", "prec_xy", "prec_z", "id", "frame", "photons")
REQUIRED_ROLES: tuple[str, ...] = ("x", "y")
COORD_ROLES: tuple[str, ...] = ("x", "y", "z")

#: Length-unit → nanometre factor (``px`` handled separately via pixel size).
_UNIT_TO_NM: dict[str, float] = {"nm": 1.0, "um": 1_000.0, "mm": 1.0e6, "m": 1.0e9}

#: Default pixel size (nm/px) when a tool stores coordinates in camera pixels
#: but no value is supplied (Picasso's typical EMCCD pixel).
DEFAULT_PIXEL_SIZE_NM: float = 130.0

# Normalised column-name synonyms per role (lower-case, units/brackets stripped).
_ROLE_SYNONYMS: dict[str, set[str]] = {
    "x": {"x", "xnm", "xpix", "xpos", "posx", "locx", "centroidx", "xmle", "xw", "xc"},
    "y": {"y", "ynm", "ypix", "ypos", "posy", "locy", "centroidy", "ymle", "yw", "yc"},
    "z": {"z", "znm", "zpix", "posz", "locz", "zmle", "zc"},
    "prec_xy": {
        "uncertainty", "uncertaintyxy", "locprecnm", "lpx", "lpy",
        "xyprecision", "precisionxy", "crlbxy", "locprec", "sigmaxy", "xprec", "yprec",
    },
    "prec_z": {
        "uncertaintyz", "locprecznm", "lpz", "precisionz", "crlbz", "zprec", "sigmaz",
    },
    "id": {
        "id", "tid", "traceid", "track", "trackid", "group", "groupindex",
        "clusterid", "cluster", "molecule", "moleculeid",
    },
    "frame": {"frame", "framenumber", "t", "time", "tim", "slice", "frameid"},
    "photons": {"photons", "phot", "intensity", "nphotons", "npho", "amplitude"},
}


@dataclass
class SpreadsheetColumn:
    name: str                # original header text
    key: str                 # normalised key (lower-case, units/brackets removed)
    unit: str | None         # "nm" | "um" | "m" | "px" | None (detected)
    values: np.ndarray       # float array (NaN for blank / non-numeric cells)
    numeric: bool            # whether the column parsed as numeric


@dataclass
class SpreadsheetTable:
    path: str
    headers: list[str]
    columns: list[SpreadsheetColumn]
    n_rows: int
    delimiter: str | None
    source_format: str       # "csv" | "tsv" | "txt" | "xlsx" | ...
    detected_tool: str = "generic"   # "thunderstorm" | "smap" | "picasso" | "generic"

    def numeric_columns(self) -> list[SpreadsheetColumn]:
        return [c for c in self.columns if c.numeric]

    def by_name(self, name: str | None) -> SpreadsheetColumn | None:
        if name is None:
            return None
        for c in self.columns:
            if c.name == name:
                return c
        return None


# ---------------------------------------------------------------------------
# Header normalisation & unit parsing
# ---------------------------------------------------------------------------

def _normalise_key(header: str) -> str:
    """Lower-case a header and strip bracketed units, spaces and punctuation.

    ``"x [nm]"`` → ``"x"`` · ``"uncertainty_xy [nm]"`` → ``"uncertaintyxy"`` ·
    ``"locprecnm"`` → ``"locprecnm"`` · ``"group index"`` → ``"groupindex"``.
    """
    h = header.strip().lower()
    h = re.sub(r"\[[^\]]*\]", "", h)     # remove [..] unit annotations
    h = re.sub(r"\([^)]*\)", "", h)      # remove (..) annotations
    return re.sub(r"[^a-z0-9]", "", h)


def _unit_token_to_name(token: str | None) -> str | None:
    if token is None:
        return None
    token = token.strip()
    if "pix" in token or token == "px":
        return "px"
    if "nm" in token:
        return "nm"
    if "um" in token or "micron" in token:
        return "um"
    if token == "m":
        return "m"
    return None


def bracket_unit(header: str) -> str | None:
    """Unit from an explicit ``[..]`` annotation only (no suffix guessing).

    Used for precision columns, whose names (``lpx``, ``xprec``) make suffix
    detection unreliable — they inherit the coordinate unit instead.
    """
    m = re.search(r"\[([^\]]+)\]", header.strip().lower().replace("µ", "u"))
    return _unit_token_to_name(m.group(1)) if m else None


def parse_unit(header: str) -> str | None:
    """Detect a length unit from a header (``nm`` / ``um`` / ``m`` / ``px``).

    Prefers an explicit bracket annotation (``x [nm]``, ``x [px]``); otherwise
    falls back to a recognised suffix (``xnm``, ``x_pix``). Returns ``None`` when
    no unit can be inferred (e.g. a bare ``x``).
    """
    h = header.strip().lower().replace("µ", "u")
    m = re.search(r"\[([^\]]+)\]", h)
    if m:
        return _unit_token_to_name(m.group(1))
    base = re.sub(r"[^a-z]", "", h)
    for suffix, unit in (("pixel", "px"), ("pix", "px"), ("px", "px"),
                         ("nm", "nm"), ("um", "um")):
        if base.endswith(suffix):
            return unit
    return None


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------

def _detect_tool(keys: set[str], headers: list[str]) -> str:
    if {"lpx", "lpy"} & keys or "netgradient" in keys:
        return "picasso"
    if "locprecnm" in keys or "xnm" in keys or "groupindex" in keys:
        return "smap"
    if any("[nm]" in h.lower() or "[photon]" in h.lower() for h in headers):
        return "thunderstorm"
    return "generic"


def _column_from_cells(name: str, cells: list[str]) -> SpreadsheetColumn:
    """Parse a column of raw string cells into a numeric (or non-numeric) column."""
    values = np.full(len(cells), np.nan, dtype=float)
    n_total = 0
    n_ok = 0
    for i, cell in enumerate(cells):
        if cell is None or cell == "":
            continue
        n_total += 1
        try:
            values[i] = float(cell)
            n_ok += 1
        except (TypeError, ValueError):
            pass
    numeric = n_total > 0 and n_ok >= 0.5 * n_total
    return SpreadsheetColumn(
        name=name, key=_normalise_key(name),
        unit=parse_unit(name), values=values, numeric=numeric,
    )


def _ordinal(n: int) -> str:
    """``1`` → ``"1st"``, ``2`` → ``"2nd"``, ``3`` → ``"3rd"``, ``4`` → ``"4th"``, …"""
    suffix = "th" if 10 <= (n % 100) <= 20 else {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def _is_header_row(cells) -> bool:
    """True when the first row looks like a **text header**, not data. A row whose
    every non-empty cell parses as a number is treated as data → no header row (so
    that first row of numbers is kept as data, and columns are named positionally)."""
    vals = [str(c).strip() for c in cells if str(c).strip() != ""]
    if not vals:
        return True
    n_numeric = 0
    for c in vals:
        try:
            float(c)
            n_numeric += 1
        except ValueError:
            pass
    return n_numeric < len(vals)          # any non-numeric cell ⇒ header


def _headers_and_data(rows: list[list], to_str) -> tuple[list[str], list[list]]:
    """Split *rows* into (headers, data rows). If the first row is not a text
    header, synthesise positional names (``"1st column"``, ``"2nd column"``, …)
    and keep the first row as data."""
    if _is_header_row(rows[0]):
        headers = [to_str(h) for h in rows[0]]
        data = rows[1:]
    else:
        n_cols = max(len(r) for r in rows)
        headers = [f"{_ordinal(i + 1)} column" for i in range(n_cols)]
        data = rows
    return headers, data


def _read_delimited(path: Path) -> tuple[list[str], list[list[str]], str]:
    """Read a delimited text file → (headers, column-major cells, delimiter)."""
    with open(path, newline="", encoding="utf-8-sig") as fh:
        sample = fh.read(8192)
    if not sample.strip():
        raise ValueError(f"'{path.name}' is empty.")
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;| ")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh, delimiter=delimiter)
        rows = [r for r in reader if any(str(c).strip() for c in r)]
    if not rows:
        raise ValueError(f"'{path.name}' has no rows.")
    headers, data = _headers_and_data(rows, lambda h: str(h).strip())
    n_cols = len(headers)
    cols: list[list[str]] = [[] for _ in range(n_cols)]
    for row in data:
        for c in range(n_cols):
            cols[c].append(row[c] if c < len(row) else "")
    return headers, cols, delimiter


def _read_excel(path: Path) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - dependency present in env
        raise ImportError("openpyxl is required to open Excel spreadsheets.") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    values = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not values:
        raise ValueError(f"'{path.name}' is empty.")
    headers, data = _headers_and_data(
        values, lambda v: str(v).strip() if v is not None else "")
    headers = [h or f"{_ordinal(i + 1)} column" for i, h in enumerate(headers)]
    n_cols = len(headers)
    cols: list[list[str]] = [[] for _ in range(n_cols)]
    for row in data:
        for c in range(n_cols):
            v = row[c] if c < len(row) else None
            cols[c].append("" if v is None else str(v))
    return headers, cols


def read_table(path: str | Path) -> SpreadsheetTable:
    """Read a spreadsheet into a :class:`SpreadsheetTable`.

    Supports ``.csv`` / ``.tsv`` / ``.txt`` (auto-detected delimiter) and
    ``.xlsx`` / ``.xlsm`` (first sheet). Raises ``ValueError`` for unsupported
    extensions or empty/headerless files.
    """
    path = Path(path)
    ext = path.suffix.lower()
    if ext in {".csv", ".tsv", ".txt"}:
        headers, cells, delimiter = _read_delimited(path)
        fmt = ext.lstrip(".")
    elif ext in {".xlsx", ".xlsm"}:
        headers, cells = _read_excel(path)
        delimiter, fmt = None, ext.lstrip(".")
    else:
        raise ValueError(
            f"Unsupported spreadsheet format '{ext}'. "
            "Supported: .csv, .tsv, .txt, .xlsx, .xlsm.")

    if not headers:
        raise ValueError(f"'{path.name}' has no header row.")
    columns = [_column_from_cells(h, cells[i]) for i, h in enumerate(headers)]
    n_rows = max((c.values.size for c in columns), default=0)
    keys = {c.key for c in columns}
    return SpreadsheetTable(
        path=str(path), headers=headers, columns=columns, n_rows=n_rows,
        delimiter=delimiter, source_format=fmt,
        detected_tool=_detect_tool(keys, headers),
    )


# ---------------------------------------------------------------------------
# Per-column statistics (cheap: vectorised numpy, no per-row Python)
# ---------------------------------------------------------------------------

#: Largest coordinate span (nm) considered a valid MINFLUX FOV (also the wrong-
#: unit guard). A localization field of view is at most ~10 µm.
SANE_MAX_SPREAD_NM: float = 1.0e5
#: MINFLUX localization coordinate constraints (used for value-based detection).
COORD_MAX_SPAN_NM: float = 1.0e4     # x/y/z spread within 10 000 nm
COORD_MIN_SPAN_NM: float = 5.0       # below this a z axis reads as flat → 2-D
#: A ``tid`` groups localizations into traces: many traces, each repeated.
TID_MIN_TRACES, TID_MAX_TRACES = 10, 10_000
TID_MIN_MEAN_COUNT, TID_MAX_MEAN_COUNT = 2.0, 1000.0
#: A monotonic non-decreasing column (this fraction of steps ≥ 0) is a time /
#: frame axis, not a scattered coordinate — the key signal distinguishing a
#: timestamp (which increases) from an x/y/z coordinate (which scatters).
MONOTONIC_FRAC: float = 0.9


@dataclass
class ColumnStats:
    """Cheap value statistics for one numeric column, used to guess its role and
    unit when the header text is missing or unrecognised."""
    n_finite: int
    is_integer: bool
    vmin: float
    vmax: float
    ptp: float                 # value range (max − min)
    median_abs_diff: float     # median |consecutive row difference|, NaN if < 2
    n_unique: int
    mean_count_per_unique: float
    frac_increasing: float     # fraction of consecutive steps that are ≥ 0 (monotonicity)


def column_stats(col: SpreadsheetColumn) -> ColumnStats:
    """Vectorised value statistics for *col* (dtype guess, range, median step,
    unique count, monotonicity). Non-numeric / empty columns return a zeroed stat."""
    v = col.values[np.isfinite(col.values)]
    if v.size == 0:
        return ColumnStats(0, False, float("nan"), float("nan"), 0.0,
                           float("nan"), 0, 0.0, 0.0)
    vmin, vmax = float(v.min()), float(v.max())
    is_int = bool(np.all(v == np.rint(v)))
    if v.size >= 2:
        d = np.diff(v)
        med = float(np.median(np.abs(d)))
        frac_inc = float(np.mean(d >= 0.0))
    else:
        med, frac_inc = float("nan"), 0.0
    n_unique = int(np.unique(v).size)
    return ColumnStats(
        n_finite=int(v.size), is_integer=is_int, vmin=vmin, vmax=vmax,
        ptp=float(vmax - vmin), median_abs_diff=med,
        n_unique=n_unique, mean_count_per_unique=v.size / max(n_unique, 1),
        frac_increasing=frac_inc,
    )


def table_stats(table: SpreadsheetTable) -> dict[str, ColumnStats]:
    """``column name → ColumnStats`` for every numeric column."""
    return {c.name: column_stats(c) for c in table.numeric_columns()}


def guess_length_unit_from_span(ptp: float) -> str:
    """Coarsest length unit (``m`` > ``mm`` > ``um`` > ``nm``) whose span stays
    within :data:`SANE_MAX_SPREAD_NM`. This disambiguates the unit from the value
    magnitude: nm data (span ~10³) stays nm, µm data (~10⁰) reads µm, and metre
    data (~10⁻⁶) reads m, because a coarser unit would overshoot 10 µm."""
    if not np.isfinite(ptp) or ptp <= 0:
        return "nm"
    for unit in ("m", "mm", "um", "nm"):
        if ptp * _UNIT_TO_NM[unit] <= SANE_MAX_SPREAD_NM:
            return unit
    return "nm"


def _is_monotonic(stats: ColumnStats) -> bool:
    """A non-decreasing axis (a timestamp / frame index increases; a coordinate
    scatters). This is the primary discriminator for the ``frame`` role."""
    return stats.n_finite >= 3 and stats.frac_increasing >= MONOTONIC_FRAC


def _is_coordinate_like(stats: ColumnStats) -> bool:
    """A **scattered** numeric column whose span fits a MINFLUX FOV (≤ 10 000 nm)
    in some length unit, above the flat-axis threshold (so a 2-D z is excluded).
    A monotonic column is a time/frame axis, not a coordinate → excluded."""
    if stats.n_finite < 3 or not np.isfinite(stats.ptp) or stats.ptp <= 0:
        return False
    if not np.isfinite(stats.median_abs_diff) or stats.median_abs_diff <= 0:
        return False                         # constant column → not a coordinate
    if _is_monotonic(stats):
        return False                         # monotonic → time/index, not a coordinate
    unit = guess_length_unit_from_span(stats.ptp)
    span_nm = stats.ptp * _UNIT_TO_NM[unit]
    return COORD_MIN_SPAN_NM <= span_nm <= COORD_MAX_SPAN_NM


def _is_tid_like(stats: ColumnStats) -> bool:
    """Integer trace id: 10–10 000 distinct values, each repeated on average."""
    return (stats.is_integer and stats.n_finite > 0
            and TID_MIN_TRACES < stats.n_unique < TID_MAX_TRACES
            and TID_MIN_MEAN_COUNT <= stats.mean_count_per_unique <= TID_MAX_MEAN_COUNT)


def time_unit_guess(stats: ColumnStats) -> str | None:
    """Time **unit** (``"s"`` / ``"ms"``) for an identified time column, from its
    median step magnitude — sub-0.1 s steps read as seconds (MINFLUX inter-loc
    intervals are ms, so a *seconds* column steps by ≪1 s), 0.1–50 read as
    milliseconds. Returns ``None`` for an **integer** column (a bare frame index,
    kept as-is) or a step outside both ranges."""
    if stats.is_integer:
        return None                          # integer counter → frame index, no s/ms
    m = stats.median_abs_diff
    if not np.isfinite(m) or m <= 0:
        return None
    if m < 0.1:
        return "s"                           # sub-0.1 s steps → seconds
    if m <= 50.0:
        return "ms"                          # 0.1–50 → milliseconds
    return None


# ---------------------------------------------------------------------------
# Auto-mapping
# ---------------------------------------------------------------------------

def guess_mapping(table: SpreadsheetTable, *, use_values: bool = False,
                  stats_map: dict[str, ColumnStats] | None = None) -> dict[str, str | None]:
    """Best-guess column name for each role, or ``None``.

    Matches each numeric column's normalised key against the per-role synonym
    sets. The first numeric column matching a role wins; a column is assigned to
    at most one role (earlier roles in :data:`ROLES` take priority).

    With ``use_values=True`` (the interactive dialog path), any **required
    localization role still unfilled after header matching** — ``x``/``y``/``z``
    (coordinates), ``id`` (tid) and ``frame`` (tim) — is guessed from the column
    **value statistics** (:func:`column_stats`): coordinate-like spans, an
    integer repeating trace id, and a millisecond-scale time step. Header matches
    always win over value guesses.
    """
    mapping: dict[str, str | None] = {role: None for role in ROLES}
    used: set[str] = set()
    for role in ROLES:
        syn = _ROLE_SYNONYMS[role]
        for col in table.columns:
            if not col.numeric or col.name in used:
                continue
            if col.key in syn:
                mapping[role] = col.name
                used.add(col.name)
                break
    if not use_values:
        return mapping

    stats_map = stats_map or table_stats(table)

    def unmapped() -> list[SpreadsheetColumn]:
        return [c for c in table.columns if c.numeric and c.name not in used]

    # tid: an integer, repeating trace id (assigned first — very distinctive).
    if mapping["id"] is None:
        for c in unmapped():
            if _is_tid_like(stats_map[c.name]):
                mapping["id"] = c.name
                used.add(c.name)
                break
    # frame / time: the first monotonic (non-decreasing) column — a timestamp /
    # frame axis increases while a coordinate scatters. Detected BEFORE coordinates
    # so a timestamp (whose span can look coordinate-like) is claimed here instead
    # of being mistaken for x.
    if mapping["frame"] is None:
        for c in unmapped():
            if _is_monotonic(stats_map[c.name]):
                mapping["frame"] = c.name
                used.add(c.name)
                break
    # x / y: the first two coordinate-like (scattered) columns, in file order.
    xy_targets = [r for r in ("x", "y") if mapping[r] is None]
    if xy_targets:
        coords = [c for c in unmapped() if _is_coordinate_like(stats_map[c.name])]
        for role, c in zip(xy_targets, coords):
            mapping[role] = c.name
            used.add(c.name)
    # z: the numeric column immediately after y, when it reads as a real out-of-plane
    # axis in the SHARED coordinate unit (x/y/z are one coordinate system). A flat
    # axis (span < COORD_MIN_SPAN_NM in that unit) stays 2-D → z unmapped.
    if mapping["z"] is None and mapping.get("x") and mapping.get("y"):
        shared = guess_length_unit_from_span(stats_map[mapping["x"]].ptp)
        cols = table.columns
        yi = next((i for i, c in enumerate(cols) if c.name == mapping["y"]), None)
        for c in (cols[yi + 1:] if yi is not None else []):
            if not c.numeric or c.name in used:
                continue
            span_nm = stats_map[c.name].ptp * _UNIT_TO_NM[shared]
            if COORD_MIN_SPAN_NM <= span_nm <= COORD_MAX_SPAN_NM:
                mapping["z"] = c.name
                used.add(c.name)
            break                     # only the column right after y is a z candidate
    return mapping


def guess_units(table: SpreadsheetTable, mapping: dict[str, str | None],
                *, stats_map: dict[str, ColumnStats] | None = None) -> dict[str, str]:
    """Per-coordinate length unit, from the column annotation, the detected tool,
    or (for a bare header with no unit) the **value magnitude**.

    Priority: an explicit header unit (``x [nm]``, ``xnm``) wins; else Picasso →
    ``px``; else :func:`guess_length_unit_from_span` picks ``nm``/``um``/``m`` from
    the coordinate's spread (so a bare ``x`` column of micrometre values reads µm).
    """
    picasso = table.detected_tool == "picasso"
    stats_map = stats_map or table_stats(table)
    # One shared value-based unit for all coordinate axes lacking a header unit,
    # taken from the axis with the LARGEST span (most reliable) — so a small
    # 3-D z inherits x/y's unit instead of being independently mis-scaled.
    spans = [stats_map[c.name].ptp for axis in COORD_ROLES
             if (c := table.by_name(mapping.get(axis))) is not None
             and not c.unit and not picasso and c.name in stats_map]
    shared = guess_length_unit_from_span(max(spans)) if spans else "nm"
    units: dict[str, str] = {}
    for axis in COORD_ROLES:
        col = table.by_name(mapping.get(axis))
        if col is None:
            units[axis] = "px" if picasso else "nm"   # unmapped → tool default
        elif col.unit:
            units[axis] = col.unit               # explicit header unit
        elif picasso:
            units[axis] = "px"
        else:
            units[axis] = shared                 # shared value-based unit
    return units


def guess_time_unit(table: SpreadsheetTable, mapping: dict[str, str | None],
                    *, stats_map: dict[str, ColumnStats] | None = None) -> str | None:
    """``"s"`` / ``"ms"`` for the mapped ``frame`` (→ tim) column from its step
    magnitude, or ``None`` when it isn't time-like (e.g. a bare frame index)."""
    col = table.by_name(mapping.get("frame"))
    if col is None:
        return None
    st = (stats_map or {}).get(col.name) or column_stats(col)
    return time_unit_guess(st)


@dataclass
class AutoImportAmbiguity:
    """Returned by :func:`auto_import` when the table can't be loaded blindly."""
    table: SpreadsheetTable
    mapping: dict[str, str | None]
    units: dict[str, str]
    reason: str
    needs_pixel_size: bool = False


def _range_corrected_units(table, mapping, units):
    """Step each coordinate's unit finer (m→mm→µm→nm) while its spread would
    exceed ``SANE_MAX_SPREAD_NM`` — catches e.g. µm values mislabelled as nm."""
    ladder = ("m", "mm", "um", "nm")
    out = dict(units)
    for axis in COORD_ROLES:
        col = table.by_name(mapping.get(axis))
        if col is None or out.get(axis) == "px":
            continue
        v = col.values[np.isfinite(col.values)]
        if v.size == 0:
            continue
        u = out[axis] if out.get(axis) in _UNIT_TO_NM else "nm"
        spread = float(np.ptp(v)) * _UNIT_TO_NM[u]
        i = ladder.index(u) if u in ladder else len(ladder) - 1
        while spread > SANE_MAX_SPREAD_NM and i < len(ladder) - 1:
            i += 1
            spread /= 1000.0
            u = ladder[i]
        out[axis] = u
    return out


def auto_import(path, *, prefs: dict | None = None, pixel_size_nm: float | None = None):
    """Try to load a spreadsheet as a dataset with **no** user interaction.

    Returns ``(dataset, None)`` on success, or ``(None, AutoImportAmbiguity)``
    when the columns are too ambiguous and the mapping dialog should be shown
    (no confident x/y, or pixel coordinates without a pixel size).
    """
    table = read_table(path)
    mapping = guess_mapping(table)
    units = guess_units(table, mapping)

    if mapping.get("x") is None or mapping.get("y") is None:
        return None, AutoImportAmbiguity(
            table, mapping, units,
            "Could not identify the x and y coordinate columns.")

    units = _range_corrected_units(table, mapping, units)

    needs_pixel = any(units.get(a) == "px" for a in COORD_ROLES if mapping.get(a))
    px = pixel_size_nm
    if needs_pixel and not px:
        px = (prefs or {}).get("data", {}).get("pixel_size_nm") or None
    if needs_pixel and not px:
        return None, AutoImportAmbiguity(
            table, mapping, units,
            "Coordinates are in camera pixels — a pixel size (nm/px) is needed.",
            needs_pixel_size=True)

    ds = build_dataset_from_mapping(
        table, mapping, units=units, pixel_size_nm=px, prefs=prefs)
    return ds, None


def representative_row_indices(n_rows: int, dense: int = 10) -> list[int]:
    """0-based row indices for a preview that spans the whole table cheaply.

    The first ``dense`` rows, then a 1-2-5 logarithmic spread (10, 20, 50, 100,
    …), and always the last row — so the user sees the start, a representative
    middle, and the end without rendering every row.
    """
    if n_rows <= 0:
        return []
    idx: set[int] = set(range(min(dense, n_rows)))
    decade = 10
    while decade < n_rows:
        for mult in (1, 2, 5):
            v = decade * mult
            if dense <= v < n_rows:
                idx.add(v)
        decade *= 10
    idx.add(n_rows - 1)
    return sorted(idx)


# ---------------------------------------------------------------------------
# Dataset construction (canonical metres)
# ---------------------------------------------------------------------------

def _to_nm(values: np.ndarray, unit: str | None, pixel_size_nm: float | None) -> np.ndarray:
    if unit == "px":
        if not pixel_size_nm or pixel_size_nm <= 0:
            raise ValueError("A positive pixel size (nm/px) is required for pixel units.")
        return values * float(pixel_size_nm)
    return values * _UNIT_TO_NM.get(unit or "nm", 1.0)


def _safe_attr_key(name: str, taken: set[str]) -> str:
    key = re.sub(r"[^0-9a-zA-Z]+", "_", name.strip()).strip("_").lower() or "col"
    if key[0].isdigit():
        key = f"c_{key}"
    base, i = key, 1
    while key in taken:
        key = f"{base}_{i}"
        i += 1
    return key


def build_dataset_from_mapping(
    table: SpreadsheetTable,
    mapping: dict[str, str | None],
    *,
    name: str | None = None,
    folder: str | None = None,
    units: dict[str, str] | None = None,
    pixel_size_nm: float | None = None,
    time_unit: str | None = None,
    prefs: dict | None = None,
):
    """Build a :class:`MinfluxDataset` from a table + a column→role mapping.

    Coordinates are converted to **metres** and stored as canonical
    ``loc_x``/``loc_y``/``loc_z`` (run through the same property/derived pipeline
    as the native loaders) so render, scatter, filters, and the precision
    analyses all work. Precision/id/frame/photons map to attributes; remaining
    numeric columns are carried through under sanitised names.

    ``time_unit`` (``"s"`` / ``"ms"`` / ``None``) rescales the ``frame`` column to
    the canonical ``tim`` in **seconds** (``"ms"`` → ÷1000; ``"s"``/``None`` kept),
    so the derived ``dt`` / ``spd`` attributes come out in sensible units.
    """
    from .dataset import build_localization_dataset

    units = units or guess_units(table, mapping)
    path = Path(table.path)
    name = name or path.name
    folder = folder if folder is not None else str(path.parent)

    def col(role: str) -> SpreadsheetColumn | None:
        return table.by_name(mapping.get(role))

    cx, cy = col("x"), col("y")
    if cx is None or cy is None:
        raise ValueError("Both x and y columns must be mapped.")
    n = int(min(cx.values.size, cy.values.size))
    if n == 0:
        raise ValueError("Mapped x/y columns are empty.")

    x_nm = _to_nm(cx.values[:n], units.get("x"), pixel_size_nm)
    y_nm = _to_nm(cy.values[:n], units.get("y"), pixel_size_nm)
    cz = col("z")
    z_nm = (_to_nm(cz.values[:n], units.get("z"), pixel_size_nm)
            if cz is not None else np.zeros(n))

    # tid from the id column (real → MINFLUX-eligible) and time from frame.
    cid = col("id")
    tid = np.asarray(cid.values[:n]) if cid is not None else None
    cframe = col("frame")
    tim = None
    if cframe is not None:
        tim = np.asarray(cframe.values[:n], dtype=float)
        if time_unit == "ms":                       # → canonical seconds
            tim = tim / 1000.0

    # Precision is in the lateral coordinate unit unless it carries an explicit
    # bracket unit (e.g. ThunderSTORM "uncertainty_xy [nm]").
    extra: dict[str, np.ndarray] = {}
    cpxy = col("prec_xy")
    loc_prec_xy = None
    if cpxy is not None:
        unit = bracket_unit(cpxy.name) or units.get("x")
        loc_prec_xy = _to_nm(cpxy.values[:n], unit, pixel_size_nm)
    cpz = col("prec_z")
    if cpz is not None:
        unit = bracket_unit(cpz.name) or units.get("z", units.get("x"))
        extra["loc_precision_z"] = _to_nm(cpz.values[:n], unit, pixel_size_nm)
    cphot = col("photons")
    if cphot is not None:
        extra["photons"] = np.asarray(cphot.values[:n], dtype=float)

    # Carry through any remaining numeric columns under sanitised keys.
    mapped_names = {mapping.get(r) for r in ROLES}
    reserved = {"loc_x", "loc_y", "loc_z", "tid", "tim",
                "loc_precision_xy", "loc_precision_z", "photons"}
    taken = set(reserved)
    for c in table.columns:
        if not c.numeric or c.name in mapped_names:
            continue
        key = _safe_attr_key(c.name, taken)
        taken.add(key)
        extra[key] = np.asarray(c.values[:n], dtype=float)

    return build_localization_dataset(
        name=name, folder=folder,
        x_nm=x_nm, y_nm=y_nm, z_nm=z_nm,
        attrs=extra, loc_precision_nm=loc_prec_xy,
        tid=tid, tim=tim, source_version="spreadsheet", prefs=prefs,
    )
